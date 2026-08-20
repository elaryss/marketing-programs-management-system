"""
Diagnostic: report which Excel rows the importer skipped and why.

Mirrors the skip logic of import_toolkit_to_supabase.py so we can answer
"which N items didn't make it into Supabase, and why?" without touching
the DB.

Uses read_only mode (works even while the file is open in Excel) and
streams rows so it's fast.

Usage:
    python scripts/diagnose_skipped_rows.py path/to/toolkit.xlsx
"""

import sys
from openpyxl import load_workbook

sys.path.insert(0, "scripts")
from import_toolkit_to_supabase import detect_columns, clean


def diagnose(xlsx_path):
    # Pass 1: load normally (non-read-only) to use the header detector,
    # but only touch row 3. We immediately switch to streaming for the body.
    # In read_only mode we can't index ws.cell directly without iterating, so
    # we instead build the header map ourselves from the first 3 rows.
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["Toolkit - Sourcing"]

    # Capture rows 1-3 (we only need row 3 for the header), plus all data rows.
    header_row = None
    data_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 3:
            header_row = row
        elif i >= 4:
            data_rows.append((i, row))

    if header_row is None:
        sys.exit("ERROR: file has fewer than 3 rows — no header found.")

    # Build a fake worksheet-like object for detect_columns. detect_columns
    # only uses ws.cell(row=3, column=c).value and ws.max_column, so we can
    # fake those with a tiny shim.
    class _Shim:
        def __init__(self, header):
            self._header = header
            self.max_column = len(header)
        def cell(self, row, column):
            class _C:
                def __init__(self, v): self.value = v
            return _C(self._header[column - 1] if column - 1 < len(self._header) else None)

    cols = detect_columns(_Shim(header_row))

    # Helper to pull a cleaned value from a row tuple by column key.
    def get(row_tuple, name):
        c = cols.get(name)
        if c is None or c - 1 >= len(row_tuple):
            return None
        return clean(row_tuple[c - 1])

    skipped_empty_rows = []   # row has *some* content but no POS and no description
    skipped_missing_bp = []   # has POS or desc, but no brand/program
    in_batch_dupes = []       # (row#, first_row#, program, pos, desc)
    parsed = 0

    seen_keys = {}  # (program, pos) -> first row#

    for excel_row, row_tuple in data_rows:
        pos = get(row_tuple, "pos_number")
        desc = get(row_tuple, "item_description")
        brand = get(row_tuple, "brand")
        program = get(row_tuple, "program")

        if not pos and not desc:
            # Only flag rows that have *any* content; ignore truly blank rows.
            if any(v not in (None, "") for v in row_tuple):
                skipped_empty_rows.append((excel_row, brand, program, row_tuple))
            continue

        if not brand or not program:
            skipped_missing_bp.append((excel_row, pos, desc, brand, program))
            continue

        key = (program, pos)
        if pos is not None and key in seen_keys:
            in_batch_dupes.append((excel_row, seen_keys[key], program, pos, desc))
            continue
        if pos is not None:
            seen_keys[key] = excel_row

        parsed += 1

    print(f"Scanned {len(data_rows)} data rows (Excel rows 4..{3 + len(data_rows)})")
    print(f"  Would be inserted:                                  {parsed}")
    print(f"  Skipped (non-blank row, no POS & no description):   {len(skipped_empty_rows)}")
    print(f"  Skipped (missing brand or program):                 {len(skipped_missing_bp)}")
    print(f"  Skipped (duplicate program+POS within file):        {len(in_batch_dupes)}")
    total_skipped = len(skipped_empty_rows) + len(skipped_missing_bp) + len(in_batch_dupes)
    print(f"  Total skipped:                                      {total_skipped}")
    print()

    if skipped_empty_rows:
        print("--- Skipped: row has content but no POS and no item_description ---")
        print(f"  {'Row':>5}  {'Brand':<20}  {'Program':<25}")
        for r, brand, program, _ in skipped_empty_rows:
            print(f"  {r:>5}  {str(brand or '—')[:20]:<20}  {str(program or '—')[:25]:<25}")
        print()

    if skipped_missing_bp:
        print("--- Skipped: missing brand or program ---")
        print(f"  {'Row':>5}  {'POS':<14}  {'Brand':<20}  {'Program':<25}  Description")
        for r, pos, desc, brand, program in skipped_missing_bp:
            print(f"  {r:>5}  {str(pos or '—')[:14]:<14}  {str(brand or '—')[:20]:<20}  {str(program or '—')[:25]:<25}  {str(desc or '')[:60]}")
        print()

    if in_batch_dupes:
        print("--- Skipped: duplicate (program, POS) — first occurrence wins ---")
        print(f"  {'Row':>5}  {'First':>5}  {'POS':<14}  {'Program':<30}  Description")
        for r, first_r, program, pos, desc in in_batch_dupes:
            print(f"  {r:>5}  {first_r:>5}  {str(pos)[:14]:<14}  {str(program)[:30]:<30}  {str(desc or '')[:60]}")
        print()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: python scripts/diagnose_skipped_rows.py <path-to-xlsx>")
    diagnose(sys.argv[1])
