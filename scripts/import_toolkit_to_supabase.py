"""
Import a Toolkit Excel file directly into Supabase.

Loads ALL lifecycle sections that exist in the F-toolkit Excel format:
- Core Elements + Sourcing + Standard Specs   → toolkit_items, paper_specs
- Pre-Buy Window Setup                        → toolkit_items (4 added fields)
- Vendor pricing (MOQ 1/2/3 tiers)            → quote_batches + vendor_quotes
- Ordering Window (Original / Revised /
  Requote / Production Phase)                 → order_snapshots
- IMS codes (ABC PO#, website description)    → toolkit_items

Auto-detects column positions from row 3 headers, so it works for
F25-F26 (113 cols), F27 (126 cols) and future layouts without manual
column mapping.

Idempotent: re-running on the same file deletes old quotes / snapshots
for affected items first, then re-inserts. toolkit_items are upserted.

Usage:
    python scripts/import_toolkit_to_supabase.py path/to/toolkit.xlsx

Reads SUPABASE_URL and SUPABASE_ANON in priority order:
  1. env vars   2. .env file   3. site/config.local.js
"""

import re
import sys
import os
import time
from pathlib import Path

import requests
from openpyxl import load_workbook


# ---------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------

def _looks_like_placeholder(s):
    if not s: return True
    up = s.upper()
    return any(x in up for x in ("PASTE", "REPLACE", "YOUR-", "YOUR_", "EXAMPLE"))


def read_config():
    url = os.environ.get("SUPABASE_URL")
    anon = os.environ.get("SUPABASE_ANON")
    if url and anon and not _looks_like_placeholder(anon):
        return url.rstrip("/"), anon

    root = Path(__file__).parent.parent
    env_path = root / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8").splitlines():
            m = re.match(r"\s*(SUPABASE_URL|SUPABASE_ANON)\s*=\s*(.+?)\s*$", line)
            if m:
                v = m.group(2).strip().strip('"').strip("'")
                if m.group(1) == "SUPABASE_URL" and not url: url = v
                if m.group(1) == "SUPABASE_ANON" and not anon: anon = v
        if url and anon and not _looks_like_placeholder(anon):
            return url.rstrip("/"), anon

    config_path = root / "site" / "config.local.js"
    if config_path.exists():
        text = config_path.read_text(encoding="utf-8")
        url_m = re.search(r'SUPABASE_URL\s*:\s*"([^"]+)"', text)
        anon_m = re.search(r'SUPABASE_ANON\s*:\s*"([^"]+)"', text)
        if url_m and anon_m and not _looks_like_placeholder(anon_m.group(1)):
            return url_m.group(1).rstrip("/"), anon_m.group(1)

    sys.exit(
        "\nERROR: Couldn't find valid Supabase credentials.\n"
        "  Tried: env vars SUPABASE_URL/SUPABASE_ANON, .env file, site/config.local.js.\n"
        "  Fix: paste your anon (publishable) key from Supabase Studio → Project Settings → API Keys\n"
        "       into site/config.local.js, replacing the placeholder."
    )


URL, KEY = read_config()
S = requests.Session()
S.headers.update({
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Content-Type": "application/json",
})


# ---------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------

def clean(v):
    if v is None: return None
    if isinstance(v, str):
        v = v.strip()
        if v in ("", "#N/A", "N/A", "n/a", "TBD", "tbd"):
            return None
    return v


def cell_num(v):
    """Coerce a value to float, or return None if not numeric / empty."""
    v = clean(v)
    if v is None: return None
    try: return float(v)
    except (ValueError, TypeError): return None


def normalize(s):
    return re.sub(r"\s+", " ", str(s).strip().lower())


COUNTRY_MAP = {
    "USA": ("US", "United States"), "US": ("US", "United States"),
    "UNITED STATES": ("US", "United States"),
    "CHINA": ("CN", "China"), "CN": ("CN", "China"),
    "MEXICO": ("MX", "Mexico"), "MX": ("MX", "Mexico"),
    "VIETNAM": ("VN", "Vietnam"), "INDIA": ("IN", "India"),
    "TAIWAN": ("TW", "Taiwan"), "ITALY": ("IT", "Italy"),
    "GERMANY": ("DE", "Germany"), "PHILIPPINES": ("PH", "Philippines"),
    "JAPAN": ("JP", "Japan"), "KOREA": ("KR", "South Korea"),
}


def resolve_country(v):
    v = clean(v)
    if not v: return None
    return COUNTRY_MAP.get(v.upper())


def normalize_category(v):
    v = clean(v)
    if not v: return "paper"
    v = v.lower()
    if "display" in v: return "display"
    if "premium" in v or "gwp" in v or "gift" in v: return "premium"
    return "paper"


# ---------------------------------------------------------------------
# Column detection — header-based, robust to layout shifts
# ---------------------------------------------------------------------

def detect_columns(ws):
    cols = {}

    def setone(key, c):
        cols.setdefault(key, c)  # first occurrence wins

    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=3, column=c).value
        if not v: continue
        h = normalize(v)

        # --- Sourcing + Core ---
        if "pre-buy program" in h:                         setone("pre_buy_program", c)
        elif "long/short" in h or h == "long/short lead":  setone("lead_time", c)
        elif "country of origin" in h:                     setone("country_of_origin", c)
        elif h.startswith("new/") and "rerun" in h:        setone("new_or_rerun", c)
        elif h == "job no." or h == "job no" or h == "job#": setone("ims_job_no", c)
        elif "sourcing vendor" in h:                       setone("sourcing_vendor", c)
        elif "sourcing responsibility" in h:               setone("sourcing_responsibility", c)
        elif h == "buyer":                                 setone("buyer", c)
        elif "rerun sku" in h:                             setone("rerun_sku_no", c)
        elif h == "pre-buy status":                        setone("pre_buy_status", c)
        elif h == "pos#" or h == "pos #":                  setone("pos_number", c)
        elif h == "brand":                                 setone("brand", c)
        elif "brand category" in h:                        setone("brand_category", c)
        elif h == "program":                               setone("program", c)
        elif "focus period" in h and "code" not in h:      setone("focus_period", c)
        elif "shipping wave" in h:                         setone("shipping_wave", c)
        elif h == "category":                              setone("category", c)
        elif "element zone" in h:                          setone("element_zone", c)
        elif h == "item type":                             setone("item_type", c)
        elif "standard/custom" in h:                       setone("standard_or_custom", c)
        elif h == "item description":                      setone("item_description", c)
        elif h == "item category":                         setone("spec_template", c)  # col AC — Standard Specs template name
        elif "item description" in h and "extended" in h:  setone("description_extended", c)
        elif "item description" in h and "sourcing" in h:  setone("sourcing_description", c)
        elif "item sub-category" in h or h == "sub-category" or h == "item sub category":
            setone("item_sub_category", c)

        # --- Standard Specs (paper) ---
        elif h == "flat size":                             setone("flat_size", c)
        elif "finished" in h and ("folded" in h or "size" in h): setone("finished_size", c)
        elif h == "material":                              setone("material", c)
        elif "coated/uncoated" in h:                       setone("coated_uncoated", c)
        elif h == "same/different":                        setone("same_different_art", c)
        elif "# of colors" in h or h == "# colors":        setone("num_colors", c)
        elif h == "coating":                               setone("coating", c)
        elif h == "finishing":                             setone("finishing", c)
        elif "collating" in h or "bundling" in h:          setone("collating", c)
        elif "unit of measure" in h or h == "uom":
            # "UOM" appears twice in this file; first is the JMS spec, second is the legacy col
            if "uom" not in cols:   cols["uom"] = c
            elif "uom_legacy" not in cols: cols["uom_legacy"] = c
        elif "quantity per uom" in h:                      setone("pack_out_qty", c)

        # --- "Pack Out" appears twice; second occurrence is the legacy column ---
        elif h == "pack out" or h == "packout":            setone("pack_out_legacy", c)

        # --- Sourcing ABC info: MOQ + price overview ---
        elif h == "moq":                                   setone("moq_overall", c)
        elif "moq in eaches" in h:                         setone("moq_overall_eaches", c)
        elif "production uom price" in h:                  setone("production_uom_price", c)
        elif h == "tariffs":                               setone("tariffs_overall", c)
        elif "uom cost without tariff" in h:               setone("uom_cost_no_tariff", c)

        # --- Vendor pricing tiers (MOQ 1/2/3) ---
        elif "moq 1 qty uom" in h:                         setone("moq1_qty_uom", c)
        elif "moq 1 qty eaches" in h:                      setone("moq1_qty_eaches", c)
        elif "moq 1 ims price" in h:                       setone("moq1_price", c)
        elif "moq 1 tariff" in h:                          setone("moq1_tariff", c)
        elif "moq 2 qty uom" in h:                         setone("moq2_qty_uom", c)
        elif "moq 2 qty eaches" in h:                      setone("moq2_qty_eaches", c)
        elif "moq 2 ims price" in h:                       setone("moq2_price", c)
        elif "moq 2 tariff" in h:                          setone("moq2_tariff", c)
        elif "moq 3 qty uom" in h:                         setone("moq3_qty_uom", c)
        elif ("moq 3 qty eaches" in h) or ("moq s qty eaches" in h):  # typo in source
            setone("moq3_qty_eaches", c)
        elif "moq 3 ims price" in h:                       setone("moq3_price", c)
        elif "moq 3 tariff" in h:                          setone("moq3_tariff", c)

        # --- Pre-Buy Window Setup ---
        elif h == "site - moq" or h == "site moq":         setone("site_moq", c)
        elif "selected production price" in h:             setone("selected_production_price", c)
        elif "portal price" in h:                          setone("portal_price", c)
        elif "estimated budget" in h:                      setone("estimated_budget_spend", c)
        elif "site setup" in h or "site slep comments" in h or "site sl3p" in h:
            setone("site_setup_comments", c)
        elif "website description" in h and "1" in h:      setone("website_description_1", c)
        elif "website description" in h and "2" in h:      setone("website_description_2", c)

        # --- Ordering Window (Original) ---
        elif "original ordered qty" in h:                  setone("original_qty", c)
        elif "budget" in h and "original" in h:            setone("original_budget", c)
        elif h.startswith("+/-") and "revised" not in h:   setone("original_plus_minus", c)
        elif "investment in moq" in h and "revised" not in h: setone("original_investment", c)
        elif "preliminary status" in h:                    setone("original_status", c)
        elif h == "abc note":                              setone("original_notes", c)

        # --- Ordering Window (Revised) ---
        elif "revised order qty" in h:                     setone("revised_qty", c)
        elif "revised budget" in h:                        setone("revised_budget", c)
        elif h.startswith("+/-") and "revised" in h:       setone("revised_plus_minus", c)
        elif "revised investment" in h:                    setone("revised_investment", c)
        elif "revsied status" in h or "revised status" in h: setone("revised_status", c)
        elif "revised - abc note" in h or "revised abc note" in h: setone("revised_notes", c)

        # --- Requoting Phase ---
        elif "qty to requote" in h and "uom" in h and "eaches" not in h:
            setone("requote_qty_uom", c)
        elif "qty to requote" in h and "eaches" in h:      setone("requote_qty_eaches", c)
        elif "production price - requoted" in h:           setone("requote_production_price", c)
        elif "sales price requoted" in h:                  setone("requote_sales_price", c)

        # --- Production Phase (Final) ---
        elif "final production qty/uom" in h or h == "final production qty/uom":
            setone("final_qty", c)
        elif "final production qty in eaches" in h:        setone("final_qty_eaches", c)
        elif "final production price" in h:                setone("final_production_price", c)
        elif "final sale price" in h or "final sales price" in h:
            setone("final_sales_price", c)
        elif "final abc budget" in h:                      setone("final_budget_spend", c)
        elif "final sales demand" in h:                    setone("final_sales_demand", c)
        elif "final sales" in h and "portal" in h:         setone("final_sales_portal", c)
        elif "final sales" in h and "actual" in h:         setone("final_sales_actual", c)
        elif "produced for inventory" in h:                setone("inventory_qty", c)
        elif "inventory cost" in h:                        setone("inventory_spend", c)
        elif "final status" in h:                          setone("final_status", c)
        elif "final abc notes" in h:                       setone("final_notes", c)
        elif "abc po#" in h or "abc po" in h:              setone("abc_po_no", c)

        # --- IMS Codes ---
        elif "brand code" in h:                            setone("brand_code", c)
        elif "program code" in h:                          setone("program_code", c)
        elif "focus period code" in h:                     setone("focus_period_code", c)
        elif "item type code" in h:                        setone("item_type_code", c)

    return cols


# ---------------------------------------------------------------------
# Supabase API helpers
# ---------------------------------------------------------------------

def api_get(path, params=None):
    r = S.get(f"{URL}/rest/v1/{path}", params=params)
    if not r.ok:
        sys.exit(f"GET {path} failed ({r.status_code}): {r.text}")
    return r.json()


def api_post(path, body, prefer="return=representation", params=None):
    headers = {"Prefer": prefer}
    r = S.post(f"{URL}/rest/v1/{path}", json=body, headers=headers, params=params)
    if not r.ok:
        sys.exit(f"POST {path} failed ({r.status_code}): {r.text}")
    return r.json() if r.text else []


def api_delete(path, params=None):
    r = S.delete(f"{URL}/rest/v1/{path}", params=params)
    if not r.ok:
        sys.exit(f"DELETE {path} failed ({r.status_code}): {r.text}")


# ---------------------------------------------------------------------
# Lookup management
# ---------------------------------------------------------------------

brand_cache = {}
program_cache = {}
item_type_cache = {}
vendor_cache = {}
country_cache = set()


def preload_lookups():
    print("Pre-loading existing lookups...")
    for b in api_get("brands", {"select": "id,name"}):
        brand_cache[b["name"]] = b["id"]
    for p in api_get("programs", {"select": "id,brand_id,name,focus_period"}):
        program_cache[(p["brand_id"], p["name"], p.get("focus_period"))] = p["id"]
    for it in api_get("item_types", {"select": "id,name"}):
        item_type_cache[it["name"]] = it["id"]
    for v in api_get("vendors", {"select": "id,name"}):
        vendor_cache[v["name"]] = v["id"]
    for c in api_get("countries", {"select": "iso_code"}):
        country_cache.add(c["iso_code"])
    print(f"  brands={len(brand_cache)} programs={len(program_cache)} "
          f"item_types={len(item_type_cache)} vendors={len(vendor_cache)} "
          f"countries={len(country_cache)}")


def ensure_brand(name, category=None):
    name = clean(name)
    if not name: return None
    if name in brand_cache: return brand_cache[name]
    row = api_post("brands", {"name": name, "category": clean(category)})[0]
    brand_cache[name] = row["id"]
    return row["id"]


def ensure_program(brand_id, name, focus_period=None, shipping_wave=None, category=None):
    name = clean(name)
    if not name or not brand_id: return None
    fp = clean(focus_period)
    key = (brand_id, name, fp)
    if key in program_cache: return program_cache[key]
    row = api_post("programs", {
        "brand_id": brand_id, "name": name, "focus_period": fp,
        "shipping_wave": clean(shipping_wave), "category": clean(category),
    })[0]
    program_cache[key] = row["id"]
    return row["id"]


def ensure_item_type(name, element_zone=None):
    name = clean(name)
    if not name: return None
    if name in item_type_cache: return item_type_cache[name]
    row = api_post("item_types", {"name": name, "element_zone": clean(element_zone)})[0]
    item_type_cache[name] = row["id"]
    return row["id"]


def ensure_vendor(name, sourcing_responsibility=None):
    name = clean(name)
    if not name: return None
    if name in vendor_cache: return vendor_cache[name]
    row = api_post("vendors", {
        "name": name, "sourcing_responsibility": clean(sourcing_responsibility),
    })[0]
    vendor_cache[name] = row["id"]
    return row["id"]


def ensure_country(iso, name):
    if iso in country_cache: return iso
    api_post("countries", {"iso_code": iso, "name": name})
    country_cache.add(iso)
    return iso


# ---------------------------------------------------------------------
# Row-level builders for each downstream table
# ---------------------------------------------------------------------

def build_quote_tier(row_data, n, cols, ws, r):
    """Return a vendor_quote tier dict for tier n (1/2/3), or None if no price."""
    price = cell_num(ws.cell(row=r, column=cols.get(f"moq{n}_price")).value) if f"moq{n}_price" in cols else None
    if price is None:
        return None
    return {
        "tier_label": f"MOQ {n}",
        "tier_number": n,
        "qty_uom":     cell_num(ws.cell(row=r, column=cols[f"moq{n}_qty_uom"]).value) if f"moq{n}_qty_uom" in cols else None,
        "qty_eaches":  cell_num(ws.cell(row=r, column=cols[f"moq{n}_qty_eaches"]).value) if f"moq{n}_qty_eaches" in cols else None,
        "moq":         cell_num(ws.cell(row=r, column=cols[f"moq{n}_qty_uom"]).value) if f"moq{n}_qty_uom" in cols else None,
        "production_price": price,
        "tariff":      cell_num(ws.cell(row=r, column=cols[f"moq{n}_tariff"]).value) if f"moq{n}_tariff" in cols else None,
    }


def build_original_snapshot(cols, ws, r):
    qty = cell_num(ws.cell(row=r, column=cols.get("original_qty")).value) if "original_qty" in cols else None
    if qty is None: return None
    return {
        "snapshot_type": "original",
        "ordered_qty": qty,
        "budget_spend": cell_num(ws.cell(row=r, column=cols.get("original_budget")).value) if "original_budget" in cols else None,
        "plus_minus_moq": cell_num(ws.cell(row=r, column=cols.get("original_plus_minus")).value) if "original_plus_minus" in cols else None,
        "investment_in_moq": cell_num(ws.cell(row=r, column=cols.get("original_investment")).value) if "original_investment" in cols else None,
        "status": clean(ws.cell(row=r, column=cols.get("original_status")).value) if "original_status" in cols else None,
        "notes":  clean(ws.cell(row=r, column=cols.get("original_notes")).value) if "original_notes" in cols else None,
    }


def build_revised_snapshot(cols, ws, r):
    qty = cell_num(ws.cell(row=r, column=cols.get("revised_qty")).value) if "revised_qty" in cols else None
    if qty is None: return None
    return {
        "snapshot_type": "revised",
        "ordered_qty": qty,
        "budget_spend": cell_num(ws.cell(row=r, column=cols.get("revised_budget")).value) if "revised_budget" in cols else None,
        "plus_minus_moq": cell_num(ws.cell(row=r, column=cols.get("revised_plus_minus")).value) if "revised_plus_minus" in cols else None,
        "investment_in_moq": cell_num(ws.cell(row=r, column=cols.get("revised_investment")).value) if "revised_investment" in cols else None,
        "status": clean(ws.cell(row=r, column=cols.get("revised_status")).value) if "revised_status" in cols else None,
        "notes":  clean(ws.cell(row=r, column=cols.get("revised_notes")).value) if "revised_notes" in cols else None,
    }


def build_requote_snapshot(cols, ws, r):
    qty_uom = cell_num(ws.cell(row=r, column=cols.get("requote_qty_uom")).value) if "requote_qty_uom" in cols else None
    price = cell_num(ws.cell(row=r, column=cols.get("requote_production_price")).value) if "requote_production_price" in cols else None
    if qty_uom is None and price is None: return None
    return {
        "snapshot_type": "requote",
        "requote_qty": qty_uom,
        "production_price": price,
        "final_sales_price": cell_num(ws.cell(row=r, column=cols.get("requote_sales_price")).value) if "requote_sales_price" in cols else None,
    }


def build_final_snapshot(cols, ws, r):
    qty = cell_num(ws.cell(row=r, column=cols.get("final_qty")).value) if "final_qty" in cols else None
    status = clean(ws.cell(row=r, column=cols.get("final_status")).value) if "final_status" in cols else None
    # Create a final row whenever the Excel has either a production qty OR a
    # Final Status. Cancel / Removed from Buy / POD / Inventory / Requoting
    # items have no production qty but still carry a terminal Final Status —
    # they MUST land in the database so variance and outcome reports can read it.
    if qty is None and not status: return None
    return {
        "snapshot_type": "final",
        "final_production_qty": qty,
        "production_price": cell_num(ws.cell(row=r, column=cols.get("final_production_price")).value) if "final_production_price" in cols else None,
        "final_sales_price": cell_num(ws.cell(row=r, column=cols.get("final_sales_price")).value) if "final_sales_price" in cols else None,
        "final_production_spend": cell_num(ws.cell(row=r, column=cols.get("final_budget_spend")).value) if "final_budget_spend" in cols else None,
        "sales_demand_qty": cell_num(ws.cell(row=r, column=cols.get("final_sales_demand")).value) if "final_sales_demand" in cols else None,
        "sales_budget_spend": cell_num(ws.cell(row=r, column=cols.get("final_sales_actual")).value) if "final_sales_actual" in cols else None,
        "inventory_qty": cell_num(ws.cell(row=r, column=cols.get("inventory_qty")).value) if "inventory_qty" in cols else None,
        "inventory_spend": cell_num(ws.cell(row=r, column=cols.get("inventory_spend")).value) if "inventory_spend" in cols else None,
        "status": status,
        "notes":  clean(ws.cell(row=r, column=cols.get("final_notes")).value) if "final_notes" in cols else None,
    }


# ---------------------------------------------------------------------
# Main import
# ---------------------------------------------------------------------

def import_file(xlsx_path):
    print(f"\nLoading {xlsx_path}...")
    wb = load_workbook(xlsx_path, data_only=True)
    if "Toolkit - Sourcing" not in wb.sheetnames:
        sys.exit("ERROR: No 'Toolkit - Sourcing' sheet in this file.")
    ws = wb["Toolkit - Sourcing"]

    cols = detect_columns(ws)
    required = ["pos_number", "brand", "program", "item_description", "category"]
    missing = [k for k in required if k not in cols]
    if missing:
        sys.exit(f"ERROR: Couldn't detect columns for: {missing}")
    print(f"Detected {len(cols)} columns. Sections covered: "
          f"core={'item_description' in cols} "
          f"specs={'flat_size' in cols} "
          f"vendor_pricing={'moq1_price' in cols} "
          f"prebuy={'selected_production_price' in cols} "
          f"original={'original_qty' in cols} "
          f"revised={'revised_qty' in cols} "
          f"final={'final_qty' in cols}")

    preload_lookups()

    def cell(r, name):
        c = cols.get(name)
        return clean(ws.cell(row=r, column=c).value) if c else None

    # Build per-row data
    print("\nParsing rows...")
    row_data_list = []
    skipped_empty = 0
    for r in range(4, ws.max_row + 1):
        pos = cell(r, "pos_number")
        desc = cell(r, "item_description")
        if not pos and not desc:
            skipped_empty += 1
            continue

        brand_name = cell(r, "brand")
        program_name = cell(r, "program")
        if not brand_name or not program_name:
            skipped_empty += 1
            continue

        brand_id = ensure_brand(brand_name, cell(r, "brand_category"))
        program_id = ensure_program(
            brand_id, program_name, cell(r, "focus_period"),
            cell(r, "shipping_wave"), cell(r, "brand_category"),
        )
        item_type_id = ensure_item_type(cell(r, "item_type"), cell(r, "element_zone"))
        vendor_name = cell(r, "sourcing_vendor")
        vendor_id = ensure_vendor(vendor_name, cell(r, "sourcing_responsibility"))

        country_resolved = resolve_country(cell(r, "country_of_origin"))
        country_code = None
        if country_resolved:
            country_code = ensure_country(country_resolved[0], country_resolved[1])

        category = normalize_category(cell(r, "category"))

        # Pack out: prefer col 43 ("Quantity per UOM"), fall back to col 50 ("Pack Out")
        pack_qty = cell_num(ws.cell(row=r, column=cols["pack_out_qty"]).value) if "pack_out_qty" in cols else None
        if pack_qty is None and "pack_out_legacy" in cols:
            pack_qty = cell_num(ws.cell(row=r, column=cols["pack_out_legacy"]).value)

        # UOM: prefer col 42 ("Unit of Measure"), fall back to col 51 ("UOM")
        uom = cell(r, "uom")
        if not uom and "uom_legacy" in cols:
            uom = clean(ws.cell(row=r, column=cols["uom_legacy"]).value)

        # Website description: combine both halves if present
        wd1 = cell(r, "website_description_1")
        wd2 = cell(r, "website_description_2")
        website = " ".join([x for x in (wd1, wd2) if x]) or None

        item_payload = {
            "program_id": program_id,
            "item_type_id": item_type_id,
            "vendor_id": vendor_id,
            "pos_number": pos,
            "shipping_wave": cell(r, "shipping_wave"),
            "category": category,
            "item_description": desc,
            "description_extended": cell(r, "description_extended"),
            "sourcing_description": cell(r, "sourcing_description"),
            "item_sub_category": cell(r, "item_sub_category"),
            "standard_or_custom": cell(r, "standard_or_custom"),
            "pack_out_qty": pack_qty,
            "uom": uom,
            "buyer": cell(r, "buyer"),
            "country_code": country_code,
            "lead_time": cell(r, "lead_time"),
            "new_or_rerun": cell(r, "new_or_rerun"),
            "pre_buy_status": cell(r, "pre_buy_status"),
            "pre_buy_program": cell(r, "pre_buy_program"),  # col A
            "spec_template":   cell(r, "spec_template"),    # col AC
            "ims_job_no": str(cell(r, "ims_job_no")) if cell(r, "ims_job_no") else None,
            # Pre-Buy Window
            "site_moq":               cell_num(ws.cell(row=r, column=cols.get("site_moq")).value) if "site_moq" in cols else None,
            "portal_price":           cell_num(ws.cell(row=r, column=cols.get("portal_price")).value) if "portal_price" in cols else None,
            "estimated_budget_spend": cell_num(ws.cell(row=r, column=cols.get("estimated_budget_spend")).value) if "estimated_budget_spend" in cols else None,
            "site_setup_comments":    cell(r, "site_setup_comments"),
            "website_description":    website,
            "abc_po_no":              str(cell(r, "abc_po_no")) if cell(r, "abc_po_no") else None,
        }

        spec_payload = None
        if category == "paper":
            nc = cell(r, "num_colors")
            spec_payload = {
                "flat_size":           cell(r, "flat_size"),
                "finished_size":       cell(r, "finished_size"),
                "material":            cell(r, "material"),
                "coated_uncoated":     cell(r, "coated_uncoated"),
                "same_different_art":  cell(r, "same_different_art"),
                "num_colors":          str(nc) if nc is not None else None,
                "coating":             cell(r, "coating"),
                "finishing":           cell(r, "finishing"),
                "collating":           cell(r, "collating"),
            }

        # Build quote tiers if any have prices
        tiers = []
        for n in (1, 2, 3):
            t = build_quote_tier(None, n, cols, ws, r)
            if t: tiers.append(t)
        quote_data = None
        if tiers and vendor_id:
            selected_price = cell_num(ws.cell(row=r, column=cols.get("selected_production_price")).value) if "selected_production_price" in cols else None
            quote_data = {"vendor_id": vendor_id, "tiers": tiers, "selected_price": selected_price}

        # Build order snapshots
        snapshots = [s for s in (
            build_original_snapshot(cols, ws, r),
            build_revised_snapshot(cols, ws, r),
            build_requote_snapshot(cols, ws, r),
            build_final_snapshot(cols, ws, r),
        ) if s]

        row_data_list.append({
            "pos": pos,
            "program_id": program_id,
            "item": item_payload,
            "spec": spec_payload,
            "quote": quote_data,
            "snapshots": snapshots,
        })

    print(f"Parsed: {len(row_data_list)} rows  |  Skipped empty: {skipped_empty}")

    # Dedupe within-file on (program_id, pos_number) — first wins
    seen, deduped, in_batch_dupes = set(), [], 0
    for rd in row_data_list:
        if rd["pos"] is not None:
            k = (rd["program_id"], rd["pos"])
            if k in seen:
                in_batch_dupes += 1
                continue
            seen.add(k)
        deduped.append(rd)
    if in_batch_dupes:
        print(f"Deduped {in_batch_dupes} within-file duplicates (same program+pos)")
    row_data_list = deduped

    if not row_data_list:
        print("Nothing to insert. Done.")
        return

    BATCH = 50
    totals = {"items": 0, "specs": 0, "quote_batches": 0, "tiers": 0, "snapshots": 0}
    t0 = time.time()

    for start in range(0, len(row_data_list), BATCH):
        batch = row_data_list[start:start + BATCH]
        item_payloads = [rd["item"] for rd in batch]

        # 1. Upsert toolkit_items (merge so existing rows get backfilled with new fields)
        inserted_items = api_post(
            "toolkit_items",
            item_payloads,
            prefer="resolution=merge-duplicates,return=representation",
            params={"on_conflict": "program_id,pos_number"},
        )
        # Key by (program_id, pos_number) — same POS can legitimately exist
        # in two different programs, so pos_number alone isn't unique.
        pos_to_id = {(it["program_id"], it["pos_number"]): it["id"]
                     for it in inserted_items if it.get("pos_number")}

        # Track which toolkit_item_ids exist in this batch so we can clear their child rows
        batch_item_ids = list(pos_to_id.values())

        # 2. Clear existing quotes / snapshots for these items so re-runs are idempotent
        if batch_item_ids:
            id_list = "(" + ",".join(batch_item_ids) + ")"
            api_delete("quote_batches", params={"toolkit_item_id": f"in.{id_list}"})
            api_delete("order_snapshots", params={"toolkit_item_id": f"in.{id_list}"})

        # 3. Insert paper_specs (upsert on toolkit_item_id which is the PK)
        spec_payloads = []
        for rd in batch:
            if rd["spec"] is None: continue
            tid = pos_to_id.get((rd["program_id"], rd["pos"]))
            if tid:
                spec_payloads.append({**rd["spec"], "toolkit_item_id": tid})
        if spec_payloads:
            api_post(
                "paper_specs",
                spec_payloads,
                prefer="resolution=merge-duplicates,return=minimal",
                params={"on_conflict": "toolkit_item_id"},
            )
            totals["specs"] += len(spec_payloads)

        # 4. Insert quote_batches → vendor_quotes
        batch_to_tiers = []  # list of (vendor_quote tier payloads, selected_price) - tier later filled with batch_id
        qb_payloads = []
        for rd in batch:
            if rd["quote"] is None: continue
            tid = pos_to_id.get((rd["program_id"], rd["pos"]))
            if not tid: continue
            qb_payloads.append({
                "toolkit_item_id": tid,
                "vendor_id":       rd["quote"]["vendor_id"],
                "source":          "excel_import",
                "status":          "selected",
            })
            batch_to_tiers.append((rd["quote"]["tiers"], rd["quote"].get("selected_price")))

        if qb_payloads:
            inserted_batches = api_post(
                "quote_batches",
                qb_payloads,
                prefer="return=representation",
            )
            totals["quote_batches"] += len(inserted_batches)

            tier_payloads = []
            select_targets = []  # (toolkit_item_id, target_tier_payload_index) — for selected_quote_id later
            for qb_row, (tiers, selected_price) in zip(inserted_batches, batch_to_tiers):
                bid = qb_row["id"]
                tid = qb_row["toolkit_item_id"]
                for tier in tiers:
                    tier_payloads.append({**tier, "quote_batch_id": bid})
                # Remember to wire selected_quote_id to MOQ 1 (default)
                if tiers:
                    select_targets.append((tid, len(tier_payloads) - len(tiers)))  # index of MOQ 1 in tier_payloads

            if tier_payloads:
                inserted_tiers = api_post(
                    "vendor_quotes",
                    tier_payloads,
                    prefer="return=representation",
                )
                totals["tiers"] += len(inserted_tiers)

                # Update toolkit_items.selected_quote_id → MOQ 1 tier
                updates = []
                for tid, idx in select_targets:
                    if idx < len(inserted_tiers):
                        updates.append((tid, inserted_tiers[idx]["id"]))
                for tid, qid in updates:
                    S.patch(
                        f"{URL}/rest/v1/toolkit_items",
                        params={"id": f"eq.{tid}"},
                        json={"selected_quote_id": qid},
                        headers={"Prefer": "return=minimal"},
                    )

        # 5. Insert order_snapshots — normalize keys so PostgREST accepts the array
        SNAP_KEYS = (
            "snapshot_type", "ordered_qty", "ordered_uom", "requote_qty",
            "final_production_qty", "sales_demand_qty", "inventory_qty",
            "budget_spend", "investment_in_moq", "production_price",
            "final_sales_price", "final_production_spend", "sales_budget_spend",
            "inventory_spend", "plus_minus_moq", "status", "notes",
        )
        snap_payloads = []
        for rd in batch:
            if not rd["snapshots"]: continue
            tid = pos_to_id.get((rd["program_id"], rd["pos"]))
            if not tid: continue
            for snap in rd["snapshots"]:
                payload = {k: snap.get(k) for k in SNAP_KEYS}
                payload["toolkit_item_id"] = tid
                snap_payloads.append(payload)
        if snap_payloads:
            api_post(
                "order_snapshots",
                snap_payloads,
                prefer="return=minimal",
            )
            totals["snapshots"] += len(snap_payloads)

        totals["items"] += len(inserted_items)
        done = start + len(batch)
        elapsed = time.time() - t0
        rate = done / elapsed if elapsed > 0 else 0
        eta = (len(row_data_list) - done) / rate if rate > 0 else 0
        print(f"  processed {done}/{len(row_data_list)}  "
              f"items+specs+quotes+snaps={totals['items']}+{totals['specs']}+{totals['tiers']}+{totals['snapshots']}  "
              f"({rate:.1f}/s, ETA {eta:.0f}s)")

    print(f"\nDone in {time.time() - t0:.1f}s:")
    print(f"  toolkit_items upserted: {totals['items']}")
    print(f"  paper_specs upserted:   {totals['specs']}")
    print(f"  quote_batches created:  {totals['quote_batches']}")
    print(f"  vendor_quotes (tiers):  {totals['tiers']}")
    print(f"  order_snapshots:        {totals['snapshots']}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit("Usage: python scripts/import_toolkit_to_supabase.py <path-to-xlsx>")
    import_file(sys.argv[1])
